"""
Reads class-roster files (.txt, .csv, .xlsx) from config.ROSTER_DIRECTORY
and parses them into (name, sid) tuples ready for
StudentRepository.insert_students().

Supported formats, all using the same convention - one or more "name"
columns/tokens followed by a numeric student id as the last value:
  .txt   "Firstname Lastname Sid" per line (whitespace separated)
  .csv   comma-separated rows, e.g. "Firstname,Lastname,Sid"
  .xlsx  spreadsheet rows, e.g. columns Name | Sid (or First | Last | Sid)

.csv/.xlsx files may optionally include a header row - it's detected and
skipped automatically (a row is treated as a header if its last cell
doesn't parse as an integer).

This is the fix for the original bug: the files under what used to be
called "exports/" (e.g. exports/DSA4042s.txt) are rosters used to populate
a course's student list, not actual exports. They now live under
config.ROSTER_DIRECTORY (data/rosters/), and real exports (CSV/Excel of
grades) go to config.EXPORT_DIRECTORY (data/exports/) via export_service.py.
"""
import csv
import os
from typing import List, Tuple

from openpyxl import load_workbook

from config import ROSTER_DIRECTORY
from logging_setup import get_logger

logger = get_logger("services.roster_import")

SUPPORTED_EXTENSIONS = (".txt", ".csv", ".xlsx")


class RosterImportError(Exception):
    pass


class RosterImportService:
    def __init__(self, roster_directory: str = ROSTER_DIRECTORY):
        self.roster_directory = roster_directory

    def list_rosters(self) -> List[str]:
        """Filenames available to pick from at course setup."""
        if not os.path.isdir(self.roster_directory):
            return []
        return sorted(
            f for f in os.listdir(self.roster_directory)
            if f.lower().endswith(SUPPORTED_EXTENSIONS)
        )

    def parse_roster_file(self, filename: str) -> List[Tuple[str, int]]:
        full_path = os.path.join(self.roster_directory, filename)
        if not os.path.isfile(full_path):
            raise RosterImportError(f"Roster file not found: {full_path}")

        ext = os.path.splitext(filename)[1].lower()
        if ext == ".txt":
            students = self._parse_txt(full_path, filename)
        elif ext == ".csv":
            students = self._parse_csv(full_path, filename)
        elif ext == ".xlsx":
            students = self._parse_excel(full_path, filename)
        else:
            raise RosterImportError(
                f"Unsupported roster file type '{ext}'. "
                f"Supported types: {', '.join(SUPPORTED_EXTENSIONS)}"
            )

        if not students:
            logger.warning("No valid student rows found in roster file %s", filename)
        return students

    # ---- format-specific parsers ----
    def _parse_txt(self, full_path: str, filename: str) -> List[Tuple[str, int]]:
        students: List[Tuple[str, int]] = []
        with open(full_path, "r", encoding="utf-8") as f:
            for line_number, line in enumerate(f, start=1):
                parts = line.strip().split()
                if len(parts) < 2:
                    logger.warning(
                        "Skipping malformed roster line %d in %s: %r",
                        line_number, filename, line.strip(),
                    )
                    continue
                *name_parts, sid_str = parts
                sid = self._try_parse_sid(sid_str)
                if sid is None:
                    logger.warning(
                        "Skipping roster line %d in %s: last token %r is not a valid Sid",
                        line_number, filename, sid_str,
                    )
                    continue
                students.append((" ".join(name_parts), sid))
        return students

    def _parse_csv(self, full_path: str, filename: str) -> List[Tuple[str, int]]:
        students: List[Tuple[str, int]] = []
        with open(full_path, "r", encoding="utf-8", newline="") as f:
            reader = csv.reader(f)
            for row_number, row in enumerate(reader, start=1):
                students_from_row = self._parse_tabular_row(row, row_number, filename)
                if students_from_row is not None:
                    students.append(students_from_row)
        return students

    def _parse_excel(self, full_path: str, filename: str) -> List[Tuple[str, int]]:
        students: List[Tuple[str, int]] = []
        try:
            workbook = load_workbook(full_path, read_only=True, data_only=True)
        except Exception as e:
            raise RosterImportError(f"Could not open Excel file '{filename}': {e}") from e

        try:
            sheet = workbook.active
            for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                cells = ["" if v is None else str(v).strip() for v in row]
                students_from_row = self._parse_tabular_row(cells, row_number, filename)
                if students_from_row is not None:
                    students.append(students_from_row)
        finally:
            workbook.close()
        return students

    # ---- shared helpers ----
    def _parse_tabular_row(self, row, row_number: int, filename: str):
        """Shared logic for CSV/Excel rows: last non-empty cell is the Sid,
        everything before it (joined) is the name. Returns None for blank
        rows or rows that look like a header."""
        cells = [str(c).strip() for c in row if str(c).strip() != ""]
        if not cells:
            return None
        if len(cells) < 2:
            logger.warning(
                "Skipping malformed roster row %d in %s: %r", row_number, filename, row,
            )
            return None

        *name_parts, sid_str = cells
        sid = self._try_parse_sid(sid_str)
        if sid is None:
            # Most likely a header row (e.g. "Name, Sid") - skip quietly on
            # row 1, warn otherwise since it's probably a real bad row.
            if row_number == 1:
                logger.info("Skipping header row in %s: %r", filename, row)
            else:
                logger.warning(
                    "Skipping roster row %d in %s: last value %r is not a valid Sid",
                    row_number, filename, sid_str,
                )
            return None

        return (" ".join(name_parts), sid)

    @staticmethod
    def _try_parse_sid(value: str):
        try:
            return int(value)
        except (TypeError, ValueError):
            return None